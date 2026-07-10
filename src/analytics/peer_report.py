from pathlib import Path
import re
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DB_PATH = Path("db/nifty100.db")
PEER_GROUPS_PATH = Path("data/raw/peer_groups.xlsx")
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "peer_comparison.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# Ten metrics required for peer ranking.
# Some document metrics are unavailable in the current source dataset,
# so equivalent available metrics are used where required.
METRICS = [
    "return_on_equity_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "interest_coverage",
    "asset_turnover",
    "operating_profit_margin_pct",
    "earnings_per_share",
    "cash_from_operations_cr",
    "composite_quality_score",
]


def clean_sheet_name(name: str) -> str:
    """Return an Excel-safe sheet name of at most 31 characters."""
    cleaned = re.sub(r"[:\\/?*\[\]]", " ", str(name)).strip()
    return cleaned[:31] or "Peer Group"


def load_latest_ratios(conn: sqlite3.Connection) -> pd.DataFrame:
    """Load the most recent financial-ratio record for each company."""
    ratios = pd.read_sql_query(
        "SELECT * FROM financial_ratios",
        conn,
    )

    ratios.columns = ratios.columns.str.strip()

    if ratios.empty:
        raise ValueError("The financial_ratios table is empty.")

    ratios["year"] = ratios["year"].astype(str)

    # Convert values such as 'Mar 2024' and 'Dec 2023' into sortable dates.
    ratios["_parsed_year"] = pd.to_datetime(
        ratios["year"],
        errors="coerce",
    )

    # When parsing fails, retain deterministic ordering.
    ratios = ratios.sort_values(
        ["company_id", "_parsed_year", "year"],
        na_position="first",
    )

    latest = ratios.groupby("company_id", as_index=False).tail(1)

    return latest.drop(columns=["_parsed_year"], errors="ignore")


def load_peer_groups() -> pd.DataFrame:
    """Load peer-group membership and benchmark information."""
    peers = pd.read_excel(PEER_GROUPS_PATH)
    peers.columns = peers.columns.str.strip()

    required = {
        "peer_group_name",
        "company_id",
        "is_benchmark",
    }

    missing = required.difference(peers.columns)

    if missing:
        raise ValueError(
            "peer_groups.xlsx is missing columns: "
            + ", ".join(sorted(missing))
        )

    peers["is_benchmark"] = (
        peers["is_benchmark"]
        .fillna(False)
        .astype(bool)
    )

    return peers


def load_latest_percentiles(
    conn: sqlite3.Connection,
    latest_ratios: pd.DataFrame,
) -> pd.DataFrame:
    """Load percentile ranks corresponding to each company's latest year."""
    percentiles = pd.read_sql_query(
        "SELECT * FROM peer_percentiles",
        conn,
    )

    if percentiles.empty:
        raise ValueError(
            "The peer_percentiles table is empty. "
            "Run: py -m src.analytics.peer"
        )

    percentiles.columns = percentiles.columns.str.strip()
    percentiles["year"] = percentiles["year"].astype(str)

    latest_keys = latest_ratios[["company_id", "year"]].drop_duplicates()

    return percentiles.merge(
        latest_keys,
        on=["company_id", "year"],
        how="inner",
    )


def prepare_peer_data() -> tuple[pd.DataFrame, list[str]]:
    """Create one latest-year, peer-ready dataset."""
    conn = sqlite3.connect(DB_PATH)

    try:
        ratios = load_latest_ratios(conn)
        peers = load_peer_groups()
        percentiles = load_latest_percentiles(conn, ratios)
    finally:
        conn.close()

    available_metrics = [
        metric for metric in METRICS
        if metric in ratios.columns
    ]

    if not available_metrics:
        raise ValueError(
            "None of the required peer metrics are present "
            "in financial_ratios."
        )

    base_columns = [
        "company_id",
        "year",
        *available_metrics,
    ]

    peer_data = peers.merge(
        ratios[base_columns],
        on="company_id",
        how="left",
    )

    percentile_wide = (
        percentiles[
            percentiles["metric"].isin(available_metrics)
        ]
        .pivot_table(
            index=[
                "company_id",
                "peer_group_name",
                "year",
            ],
            columns="metric",
            values="percentile_rank",
            aggfunc="first",
        )
        .reset_index()
    )

    percentile_wide.columns = [
        (
            f"{column}_percentile"
            if column in available_metrics
            else column
        )
        for column in percentile_wide.columns
    ]

    peer_data["year"] = peer_data["year"].astype(str)

    peer_data = peer_data.merge(
        percentile_wide,
        on=[
            "company_id",
            "peer_group_name",
            "year",
        ],
        how="left",
    )

    # The source dataset uses company_id as the stock/company identifier.
    # Retain a separate display column required by the report specification.
    peer_data.insert(
        1,
        "company_name",
        peer_data["company_id"],
    )

    return peer_data, available_metrics


def add_median_row(
    group: pd.DataFrame,
    metrics: list[str],
) -> pd.DataFrame:
    """Append a peer-group median summary row."""
    percentile_columns = [
        f"{metric}_percentile"
        for metric in metrics
        if f"{metric}_percentile" in group.columns
    ]

    summary = {
        column: None
        for column in group.columns
    }

    summary["company_id"] = "PEER MEDIAN"
    summary["company_name"] = "Peer Group Median"
    summary["peer_group_name"] = group[
        "peer_group_name"
    ].iloc[0]
    summary["is_benchmark"] = False
    summary["year"] = "Summary"

    numeric_columns = metrics + percentile_columns

    for column in numeric_columns:
        if column in group.columns:
            values = pd.to_numeric(
                group[column],
                errors="coerce",
            )
            summary[column] = values.median()

    return pd.concat(
        [group, pd.DataFrame([summary])],
        ignore_index=True,
    )


def export_peer_comparison() -> None:
    """Generate and format the peer-comparison workbook."""
    peer_data, metrics = prepare_peer_data()

    peer_groups = sorted(
        peer_data["peer_group_name"]
        .dropna()
        .astype(str)
        .unique()
    )

    print(f"Peer groups found: {len(peer_groups)}")

    if len(peer_groups) != 11:
        print(
            "Warning: the source file contains "
            f"{len(peer_groups)} peer groups; "
            "the Sprint 3 exit criterion expects 11."
        )

    percentile_columns = [
        f"{metric}_percentile"
        for metric in metrics
        if f"{metric}_percentile" in peer_data.columns
    ]

    report_columns = [
        "company_id",
        "company_name",
        "peer_group_name",
        "year",
        "is_benchmark",
        *metrics,
        *percentile_columns,
    ]

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        for peer_group in peer_groups:
            group = peer_data[
                peer_data["peer_group_name"] == peer_group
            ].copy()

            # Keep only latest-year company records.
            group = group.drop_duplicates(
                subset=["company_id"],
                keep="last",
            )

            group = group[
                [
                    column
                    for column in report_columns
                    if column in group.columns
                ]
            ]

            group = group.sort_values(
                by=[
                    "is_benchmark",
                    "company_id",
                ],
                ascending=[False, True],
            )

            group = add_median_row(group, metrics)
            
            group = group.drop(columns=["peer_group_name"], errors="ignore")

            group.to_excel(
                writer,
                sheet_name=clean_sheet_name(peer_group),
                index=False,
            )

    format_workbook(percentile_columns)

    print("=" * 60)
    print("Peer comparison workbook generated.")
    print(f"Sheets created: {len(peer_groups)}")
    print(f"Saved to: {OUTPUT_FILE}")
    print("=" * 60)


def format_workbook(percentile_columns: list[str]) -> None:
    """Apply report formatting and percentile colour coding."""
    workbook = load_workbook(OUTPUT_FILE)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    benchmark_fill = PatternFill(
        fill_type="solid",
        fgColor="FFD966",
    )
    median_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    high_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )
    middle_fill = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )
    low_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_map = {}

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            header_map[cell.value] = cell.column

        benchmark_column = header_map.get("is_benchmark")
        company_column = header_map.get("company_id")

        percentile_indices = [
            header_map[column]
            for column in percentile_columns
            if column in header_map
        ]

        for row in range(2, worksheet.max_row + 1):
            company_value = (
                worksheet.cell(
                    row=row,
                    column=company_column,
                ).value
                if company_column
                else None
            )

            is_median = company_value == "PEER MEDIAN"

            is_benchmark = False

            if benchmark_column:
                raw_value = worksheet.cell(
                    row=row,
                    column=benchmark_column,
                ).value

                is_benchmark = raw_value in (
                    True,
                    1,
                    "True",
                    "TRUE",
                )

            if is_median:
                for cell in worksheet[row]:
                    cell.fill = median_fill
                    cell.font = Font(bold=True)

            elif is_benchmark:
                for cell in worksheet[row]:
                    cell.fill = benchmark_fill
                    cell.font = Font(bold=True)

            # Do not overwrite benchmark or median row colouring.
            if not is_median and not is_benchmark:
                for column_index in percentile_indices:
                    cell = worksheet.cell(
                        row=row,
                        column=column_index,
                    )

                    try:
                        value = float(cell.value)
                    except (TypeError, ValueError):
                        continue

                    if value >= 0.75:
                        cell.fill = high_fill
                    elif value <= 0.25:
                        cell.fill = low_fill
                    else:
                        cell.fill = middle_fill

                    cell.number_format = "0.0%"

        # Format numeric values and improve readability.
        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):
            header = worksheet.cell(
                row=1,
                column=column_index,
            ).value

            if header in percentile_columns:
                for row in range(
                    2,
                    worksheet.max_row + 1,
                ):
                    worksheet.cell(
                        row=row,
                        column=column_index,
                    ).number_format = "0.0%"
            elif header not in {
                "company_id",
                "company_name",
                "year",
                "is_benchmark",
            }:
                for row in range(
                    2,
                    worksheet.max_row + 1,
                ):
                    worksheet.cell(
                        row=row,
                        column=column_index,
                    ).number_format = "0.00"

            max_length = len(str(header or ""))

            for row in range(
                2,
                worksheet.max_row + 1,
            ):
                value = worksheet.cell(
                    row=row,
                    column=column_index,
                ).value
                max_length = max(
                    max_length,
                    len(str(value or "")),
                )

            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(max_length + 2, 28)

        worksheet.row_dimensions[1].height = 34

    workbook.save(OUTPUT_FILE)


if __name__ == "__main__":
    export_peer_comparison()